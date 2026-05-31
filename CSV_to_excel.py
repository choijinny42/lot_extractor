import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO

def format_excel(df, sheet_name="extracted data"):
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name=sheet_name)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    ws.title = sheet_name

    # wrap all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # adjust column widths
    MAX_WIDTH = 50
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, MAX_WIDTH)

    # make into table
    table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="Table1", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # save back to BytesIO
    formatted = BytesIO()
    wb.save(formatted)
    formatted.seek(0)
    return formatted